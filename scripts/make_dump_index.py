#!/usr/bin/env python3
"""
Generate a master index Excel file across all courses in a target dump.

Walks every course folder in the dump root, reads course_metadata.json,
and writes a single index.xlsx to the dump root so non-technical staff
can quickly identify and navigate all courses.

Usage:
  python scripts/make_dump_index.py
  python scripts/make_dump_index.py --dump-root export/data/target_dump
  python scripts/make_dump_index.py --out export/data/target_dump/index.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

THIS_FILE = Path(__file__).resolve()
REPO_ROOT = THIS_FILE.parents[1]


def _read_json(path: Path) -> dict | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _style_header(ws, color: str = "1F4E79") -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def build_index(dump_root: Path, out_path: Path) -> None:
    course_dirs = sorted(
        (d for d in dump_root.iterdir() if d.is_dir() and d.name.isdigit()),
        key=lambda d: int(d.name),
    )

    if not course_dirs:
        sys.exit(f"No numeric course folders found in {dump_root}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Index"
    ws.freeze_panes = "A2"

    headers = [
        "Course ID",
        "Course Code",
        "SIS Course ID",
        "Course Name",
        "Status",
        "Blueprint",
        "Folder",
        "Manifest",
    ]
    ws.append(headers)
    _style_header(ws)

    missing_meta = []
    for course_dir in course_dirs:
        meta_path = course_dir / "course" / "course_metadata.json"
        meta = _read_json(meta_path)
        if not meta:
            missing_meta.append(course_dir.name)
            meta = {}

        course_id   = meta.get("id") or course_dir.name
        course_code = meta.get("course_code") or ""
        sis_id      = meta.get("sis_course_id") or ""
        name        = meta.get("name") or ""
        status      = (meta.get("workflow_state") or "").replace("_", " ").title()
        blueprint   = "Yes" if meta.get("is_blueprint") else "No"

        manifest_path = course_dir / "manifest.xlsx"
        manifest_rel  = str(manifest_path.relative_to(dump_root)) if manifest_path.exists() else ""

        row_idx = ws.max_row + 1
        ws.append([
            course_id,
            course_code,
            sis_id,
            name,
            status,
            blueprint,
            str(course_dir.relative_to(dump_root)),
            manifest_rel,
        ])

        # Make manifest cell a hyperlink if the file exists
        if manifest_path.exists():
            cell = ws.cell(row=row_idx, column=8)
            cell.hyperlink = manifest_path.as_uri()
            cell.font = Font(color="0563C1", underline="single")

    _autofit(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"Index written: {out_path}")
    print(f"  Courses indexed: {len(course_dirs)}")
    if missing_meta:
        print(f"  WARNING: missing course_metadata.json for {len(missing_meta)} courses: {missing_meta[:10]}")


def main() -> int:
    p = argparse.ArgumentParser(description="Generate a master index across all dump courses")
    p.add_argument("--dump-root", type=Path, default=Path("export/data/target_dump"),
                   help="Root of the target dump (default: export/data/target_dump)")
    p.add_argument("--out", type=Path, default=None,
                   help="Output path (default: <dump-root>/index.xlsx)")
    args = p.parse_args()

    dump_root = args.dump_root.resolve()
    if not dump_root.is_dir():
        p.error(f"Dump root not found: {dump_root}")

    out_path = args.out or (dump_root / "index.xlsx")
    build_index(dump_root, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
