#!/usr/bin/env python3
"""
Batch exporter for BAC BPM courses flagged as ready.

Usage examples:
  python scripts/batch_export_bac.py --excel docs/data/batch_imports.xlsx --dry-run
  python scripts/batch_export_bac.py --excel docs/data/batch_imports.xlsx \
      --export-root export/data --batch-size 3 --steps pages assignments

Requires openpyxl: pip install openpyxl
"""
from __future__ import annotations

import argparse
import subprocess
import sys
from itertools import islice
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

def append_course_id(output_file: Path, course_id: int) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with output_file.open("a", encoding="utf-8") as fh:
        fh.write(f"{course_id}\n")


def normalize_bpm(value: str | None) -> str:
    if not value:
        return ""
    return value.replace("\xa0", "").strip()


def is_ready(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"1", "yes", "y", "true"}


def iter_ready_courses(excel_path: Path, sheet_name: str = "BPMs") -> list[tuple[str, int]]:
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet '{sheet_name}' not found in {excel_path}")
    ws = wb[sheet_name]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    required = {"BPM", "Course ID", "Ready for Migration"}
    missing = required - idx.keys()
    if missing:
        raise ValueError(f"missing expected columns: {', '.join(sorted(missing))}")

    rows: list[tuple[str, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bpm_raw = normalize_bpm(row[idx["BPM"]])
        if not bpm_raw.startswith("BAC"):
            continue
        if not is_ready(row[idx["Ready for Migration"]]):
            continue
        course_raw = row[idx["Course ID"]]
        if course_raw is None:
            continue
        try:
            course_id = int(str(course_raw).strip())
        except ValueError:
            print(f"Skipping row with non-numeric Course ID: {course_raw!r}", file=sys.stderr)
            continue
        rows.append((bpm_raw, course_id))
    return rows


def batched(iterable: Iterable[tuple[str, int]], batch_size: int) -> Iterable[list[tuple[str, int]]]:
    it = iter(iterable)
    while True:
        batch = list(islice(it, batch_size))
        if not batch:
            break
        yield batch


def run_export(course_id: int, export_root: Path, steps: list[str], dry_run: bool) -> None:
    cmd = [
        sys.executable,
        "scripts/run_export.py",
        "--course-id",
        str(course_id),
        "--export-root",
        str(export_root),
    ]
    if steps:
        cmd += ["--steps", *steps]
    print(" ".join(cmd) if dry_run else f"Running: {' '.join(cmd)}")
    if dry_run:
        return
    subprocess.run(cmd, check=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run exports for BAC BPM courses marked ready.")
    parser.add_argument("--excel", type=Path, required=True, help="Path to docs/data/batch_imports.xlsx")
    parser.add_argument("--export-root", type=Path, default=Path("export/data"), help="Export root directory")
    parser.add_argument("--steps", nargs="+", default=[], help="Optional subset of export steps")
    parser.add_argument("--batch-size", type=int, default=5, help="Number of courses per batch")
    parser.add_argument("--limit", type=int, default=None, help="Process only the first N courses")
    parser.add_argument("--dry-run", action="store_true", help="Print commands without running them")
 
    parser.add_argument(
        "--record-path",
        type=Path,
        default=Path("docs/exported_course_ids.txt"),
        help="Where to append exported course IDs (one per line)",
    )
    parser.add_argument(
        "--record-reset",
        action="store_true",
        help="Clear the record file before starting a new batch",
    )

    args = parser.parse_args()
    
    if args.record_path and args.record_reset and args.record_path.exists():
        args.record_path.unlink()


    courses = iter_ready_courses(args.excel)
    if args.limit is not None:
        courses = courses[: args.limit]

    print(f"Found {len(courses)} ready BAC courses.")
    for batch_num, batch in enumerate(batched(courses, args.batch_size), start=1):
        print(f"\nBatch {batch_num}: {[cid for _, cid in batch]}")
        for bpm, course_id in batch:
            print(f"  {bpm} -> {course_id}")
            try:
                run_export(course_id, args.export_root, args.steps, args.dry_run)
                if args.record_path and not args.dry_run:
                    append_course_id(args.record_path, course_id)

            except subprocess.CalledProcessError as exc:
                print(f"Export failed for {course_id}: {exc}", file=sys.stderr)
                if not args.dry_run:
                    return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
