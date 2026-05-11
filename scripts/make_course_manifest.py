#!/usr/bin/env python3
"""
Generate a human-readable Excel manifest for an exported Canvas course.

Reads from the local export directory produced by run_export.py and writes
a multi-sheet workbook so non-technical staff can review course content.

Usage:
  python scripts/make_course_manifest.py --course-dir export/data/12345
  python scripts/make_course_manifest.py --course-dir export/data/12345 --out my_manifest.xlsx

Requires openpyxl: pip install openpyxl
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_date(raw: str | None) -> str:
    if not raw:
        return ""
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d")
    except ValueError:
        return raw


def _published(value: Any) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return ""


def _read_json(path: Path) -> dict | list | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _clean_filename(raw: str) -> str:
    """Strip the Canvas timestamp prefix (e.g. '1734461109_43__') from filenames."""
    import re
    return re.sub(r"^\d+_\d+__", "", raw)


def _style_header_row(ws, header_fill: str = "1F4E79") -> None:
    fill = PatternFill("solid", fgColor=header_fill)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def _freeze(ws) -> None:
    ws.freeze_panes = ws["A2"]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_summary(wb: Workbook, course_dir: Path, counts: dict[str, int]) -> None:
    meta_path = course_dir / "course" / "course_metadata.json"
    meta: dict = _read_json(meta_path) or {}

    ws = wb.active
    ws.title = "Summary"

    ws.append(["Field", "Value"])
    _style_header_row(ws)

    rows = [
        ("Course ID",       meta.get("id", course_dir.name)),
        ("Course Name",     meta.get("name", "")),
        ("Course Code",     meta.get("course_code", "")),
        ("SIS Course ID",   meta.get("sis_course_id", "")),
        ("Status",          meta.get("workflow_state", "").replace("_", " ").title()),
        ("Is Blueprint",    "Yes" if meta.get("is_blueprint") else "No"),
        ("Manifest Date",   datetime.now().strftime("%Y-%m-%d")),
        ("", ""),
        ("Content Type",    "Count"),
    ]
    for r in rows:
        ws.append(r)

    for label, key in [
        ("Modules",       "modules"),
        ("Pages",         "pages"),
        ("Assignments",   "assignments"),
        ("Quizzes",       "quizzes"),
        ("Discussions",   "discussions"),
        ("Announcements", "announcements"),
        ("Files",         "files"),
    ]:
        ws.append([label, counts.get(key, 0)])

    _autofit(ws)


def _sheet_modules(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Modules")
    ws.append(["Position", "Module Name", "Published", "Items"])
    _style_header_row(ws)
    _freeze(ws)

    modules_json = course_dir / "modules" / "modules.json"
    modules: list = _read_json(modules_json) or []

    for mod in modules:
        ws.append([
            mod.get("position", ""),
            mod.get("name", ""),
            _published(mod.get("published")),
            mod.get("items_count") or "",
        ])

    _autofit(ws)
    return len(modules)


def _sheet_pages(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Pages")
    ws.append(["#", "Title", "Published", "Last Updated"])
    _style_header_row(ws)
    _freeze(ws)

    pages_dir = course_dir / "pages"
    count = 0
    for page_dir in sorted(pages_dir.iterdir()) if pages_dir.exists() else []:
        if not page_dir.is_dir():
            continue
        meta: dict = _read_json(page_dir / "page_metadata.json") or {}
        if not meta:
            continue
        count += 1
        ws.append([
            meta.get("position", count),
            meta.get("title", ""),
            _published(meta.get("published")),
            _fmt_date(meta.get("updated_at")),
        ])

    _autofit(ws)
    return count


def _sheet_assignments(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Assignments")
    ws.append(["#", "Name", "Points", "Due Date", "Submission Type", "Grading", "Published"])
    _style_header_row(ws)
    _freeze(ws)

    asgn_dir = course_dir / "assignments"
    count = 0
    for asgn_dir_item in sorted(asgn_dir.iterdir()) if asgn_dir.exists() else []:
        if not asgn_dir_item.is_dir():
            continue
        meta: dict = _read_json(asgn_dir_item / "assignment_metadata.json") or {}
        if not meta:
            continue
        count += 1
        sub_types = meta.get("submission_types") or []
        ws.append([
            count,
            meta.get("name", ""),
            meta.get("points_possible", ""),
            _fmt_date(meta.get("due_at")),
            ", ".join(sub_types).replace("_", " "),
            (meta.get("grading_type") or "").replace("_", " "),
            _published(meta.get("published")),
        ])

    _autofit(ws)
    return count


def _sheet_quizzes(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Quizzes")
    ws.append(["#", "Title", "Type", "Points", "Questions", "Attempts", "Due Date", "Published"])
    _style_header_row(ws)
    _freeze(ws)

    quiz_dir = course_dir / "quizzes"
    count = 0
    for qdir in sorted(quiz_dir.iterdir()) if quiz_dir.exists() else []:
        if not qdir.is_dir():
            continue
        meta: dict = _read_json(qdir / "quiz_metadata.json") or {}
        if not meta:
            continue
        count += 1
        attempts = meta.get("allowed_attempts", "")
        ws.append([
            count,
            meta.get("title", ""),
            (meta.get("quiz_type") or "").replace("_", " "),
            meta.get("points_possible", ""),
            meta.get("question_count", ""),
            "Unlimited" if attempts == -1 else attempts,
            _fmt_date(meta.get("due_at")),
            _published(meta.get("published")),
        ])

    _autofit(ws)
    return count


def _sheet_discussions(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Discussions")
    ws.append(["#", "Title", "Type", "Graded", "Published", "Posted"])
    _style_header_row(ws)
    _freeze(ws)

    disc_dir = course_dir / "discussions"
    count = 0
    for ddir in sorted(disc_dir.iterdir()) if disc_dir.exists() else []:
        if not ddir.is_dir():
            continue
        meta: dict = _read_json(ddir / "discussion_metadata.json") or {}
        if not meta:
            continue
        count += 1
        ws.append([
            count,
            meta.get("title", ""),
            (meta.get("discussion_type") or "").replace("_", " "),
            "Yes" if meta.get("assignment") else "No",
            _published(meta.get("published")),
            _fmt_date(meta.get("posted_at")),
        ])

    _autofit(ws)
    return count


def _sheet_announcements(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Announcements")
    ws.append(["#", "Title", "Published", "Posted"])
    _style_header_row(ws)
    _freeze(ws)

    ann_dir = course_dir / "announcements"
    count = 0
    for adir in sorted(ann_dir.iterdir()) if ann_dir.exists() else []:
        if not adir.is_dir():
            continue
        meta: dict = _read_json(adir / "announcement_metadata.json") or {}
        if not meta:
            continue
        count += 1
        ws.append([
            count,
            meta.get("title", ""),
            _published(meta.get("published")),
            _fmt_date(meta.get("posted_at")),
        ])

    _autofit(ws)
    return count


def _sheet_files(wb: Workbook, course_dir: Path) -> int:
    ws = wb.create_sheet("Files")
    ws.append(["#", "File Name", "Folder", "File Type", "Last Updated"])
    _style_header_row(ws)
    _freeze(ws)

    files_dir = course_dir / "files"
    count = 0
    for meta_path in sorted(files_dir.rglob("*.metadata.json")) if files_dir.exists() else []:
        meta: dict = _read_json(meta_path) or {}
        if not meta:
            continue
        count += 1
        raw_filename = meta.get("filename") or meta_path.stem.removesuffix(".metadata")
        ws.append([
            count,
            _clean_filename(raw_filename),
            meta.get("folder_path", ""),
            meta.get("content_type", ""),
            _fmt_date(meta.get("updated_at")),
        ])

    _autofit(ws)
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_manifest(course_dir: Path, out_path: Path) -> None:
    if not course_dir.is_dir():
        sys.exit(f"Course directory not found: {course_dir}")

    wb = Workbook()

    counts: dict[str, int] = {}
    counts["modules"]       = _sheet_modules(wb, course_dir)
    counts["pages"]         = _sheet_pages(wb, course_dir)
    counts["assignments"]   = _sheet_assignments(wb, course_dir)
    counts["quizzes"]       = _sheet_quizzes(wb, course_dir)
    counts["discussions"]   = _sheet_discussions(wb, course_dir)
    counts["announcements"] = _sheet_announcements(wb, course_dir)
    counts["files"]         = _sheet_files(wb, course_dir)

    # Summary sheet must be built last (needs counts) but placed first
    _sheet_summary(wb, course_dir, counts)
    # Move Summary to front
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Manifest written: {out_path}")
    for key, n in counts.items():
        print(f"  {key:<15} {n}")


def main() -> int:
    p = argparse.ArgumentParser(description="Generate an Excel manifest for an exported Canvas course")
    p.add_argument("--course-dir", type=Path, required=True,
                   help="Path to exported course directory (e.g. export/data/12345)")
    p.add_argument("--out", type=Path, default=None,
                   help="Output path for the manifest (default: <course-dir>/manifest.xlsx)")
    args = p.parse_args()

    course_dir = args.course_dir.resolve()
    out_path = args.out or (course_dir / "manifest.xlsx")

    build_manifest(course_dir, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
